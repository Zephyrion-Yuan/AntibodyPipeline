import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from typing import Dict, List
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import os
from shutil import copyfile

class PlateSheet:
    def __init__(self, name: str, data: pd.DataFrame, well_volumes: dict = None, default_volume: float = 800.0):
        self.name = name
        self.data = data
        self.number_to_value = {}
        self.value_to_numbers = {}
        self._build_mappings()
        # 新增体积属性，初始化
        if well_volumes is not None:
            self.well_volumes = dict(well_volumes)  # 传dict则拷贝
        elif default_volume is not None:
            self.well_volumes = {i: default_volume for i in range(1, 97)}
        else:
            self.well_volumes = {}  # 空字典
        self.full_paired = {i: False for i in range(1, 97)}

    def _build_mappings(self):
        content = self.data.iloc[1:, 1:]
        idx = 1
        for col in content.columns:
            for row in content.index:
                val = content.loc[row, col].lower()
                self.number_to_value[idx] = val
                if val not in self.value_to_numbers:
                    self.value_to_numbers[val] = []
                self.value_to_numbers[val].append(idx)
                idx += 1

    def get_numbered_values(self) -> Dict[int, str]:
        return self.number_to_value

    def get_value_by_number(self, number: int) -> str:
        return self.number_to_value.get(number, None)

    def get_number_by_value(self, value: str) -> list:
        return self.value_to_numbers.get(value, [])
    
    def get_highlight_positions(self, positions: list) -> set:
        """
        根据一组编号，返回需要高亮的(row, col)坐标集合，便于Excel写入。
        返回的row, col均为1-based，直接用于openpyxl.cell(row=, column=)
        """
        coords = set()
        for num in positions:
            col = ((num - 1) // 8) + 2
            row = ((num - 1) % 8) + 2
            coords.add((row, col))
        return coords
    
    def is_accessible(self, position: int, volume: float) -> bool:
        """判断某孔是否可以取用指定体积。"""
        well_vol = self.well_volumes.get(position, None)
        if well_vol is None:
            # 未设置体积，默认不可访问
            return False
        return volume <= well_vol

    def take_volume(self, position: int, volume: float) -> bool:
        """实际取用体积，体积够则扣减，返回True，否则返回False。"""
        if self.is_accessible(position, volume):
            self.well_volumes[position] -= volume
            return True
        return False

    def set_well_volume(self, position: int, volume: float):
        """设置某孔体积。"""
        self.well_volumes[position] = volume
        return

    def get_well_volume(self, position: int) -> float:
        """获取某孔体积。"""
        return self.well_volumes.get(position, None)
    
    # def get_is_full_paired(self, position: int) -> bool:
    #     """判断是否已经全配对"""
    #     return self.full_paired.get(position, None)
    

class PlateWorkbook:
    def __init__(self):
        self.plates = {}
        self.plates_num = 0

    # def read_from_excel(self, file_path: str):
    #     xls = pd.read_excel(file_path, sheet_name=None, engine='openpyxl',header=None, keep_default_na=False, na_values=['nan'])
    #     for sheet_name, df in xls.items():
    #         # if sheet_name.lower().startswith('plate'):
    #         #     self.plates[sheet_name] = PlateSheet(sheet_name, df.fillna(""))
    #         self.plates[sheet_name] = PlateSheet(sheet_name, df.fillna(""))
    #     return

    def read_from_excel(self, file_path: str):
        xls = pd.read_excel(
            file_path,
            sheet_name=None,
            engine='openpyxl',
            header=None,
            keep_default_na=False,
            na_values=['nan']
        )
        for sheet_name, df in xls.items():
            # 填充空值为字符串空（统一判断标准）
            df_filled = df.fillna("")
            # 提取除第一行（索引0）和第一列（索引0）之外的区域
            # iloc[1:, 1:] 表示：行从索引1到末尾，列从索引1到末尾
            content_area = df_filled.iloc[1:, 1:]
            # 判断该区域是否为空：
            # 1. 区域本身无数据（如只有1行或1列）
            # 2. 区域内所有单元格都是空字符串
            if content_area.empty or (content_area == "").all().all():
                continue  # 为空则跳过
            # 非空则保留该sheet
            self.plates[sheet_name] = PlateSheet(sheet_name, df_filled)
            self.plates_num += 1
        return

    def write_to_excel(self, output_path: str):
        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet
        for plate_name, plate_sheet in self.plates.items():
            ws = wb.create_sheet(title=plate_name)
            # write headers
            headers = plate_sheet.data.columns.tolist()
            ws.append(headers)
            for i in range(1, plate_sheet.data.shape[0]):
                row_data = [plate_sheet.data.iloc[i, j] for j in range(plate_sheet.data.shape[1])]
                ws.append(row_data)
        wb.save(output_path)
        return

    def get_all_numbered_data(self) -> Dict[str, Dict[int, str]]:
        """
        Returns:
            Dict where keys are plate names and values are dictionaries mapping 1-96 to data
        """
        return {name: plate.get_numbered_values() for name, plate in self.plates.items()}
    
    def highlight_and_save(
        self, 
        file_path: str, 
        plate_highlight_dict: dict,  # dict: plate_name -> set(编号int)
        output_dir: str = None,
        color="FFFF00"
    ):
        """
        根据plate_highlight_dict中每个plate要高亮的编号，生成带高亮的新Excel。
        - file_path: 原始plate Excel文件
        - plate_highlight_dict: {plate_name: set([编号, ...]), ...}
        - output_dir: 输出文件夹，默认同目录
        - color: 高亮颜色（默认为黄色）
        """
        if output_dir is None:
            output_dir = os.path.dirname(file_path)
        output_file = os.path.join(output_dir, f"highlighted_{os.path.basename(file_path)}")

        # 拷贝并打开
        copyfile(file_path, output_file)
        wb = load_workbook(output_file)
        fill = PatternFill(fill_type="solid", fgColor=color)

        # 遍历每个板
        for plate_name, positions in plate_highlight_dict.items():
            sheet_name = plate_name if plate_name in wb.sheetnames else f"plate{plate_name}"
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # 用PlateSheet的接口计算所有cell
                coords = self.plates[plate_name].get_highlight_positions(positions)
                for row, col in coords:
                    ws.cell(row=row, column=col).fill = fill

        wb.save(output_file)
        return
        
    def add_plate_from_list(self, data_list, plate_name):
        assert len(data_list) <= 96
        rows = [plate_name, 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        cols = [plate_name] + [str(i) for i in range(1, 13)]
        df = pd.DataFrame('', index=rows, columns=cols)
        for i, letter in enumerate(rows[1:], 1):
            df.iloc[i, 0] = letter
        for j, number in enumerate(cols[1:], 1):
            df.iloc[0, j] = number
        idx = 0
        for col in range(1, 13):
            for row in range(1, 9):
                df.iloc[row, col] = data_list[idx] if idx < len(data_list) else ""
                idx += 1
        self.plates[plate_name] = PlateSheet(plate_name, df)
        return
    
'''
from plate_structure import PlateWorkbook

# 1. 创建实例
wb = PlateWorkbook()

# 2. 读取Excel文件（包含 plate1, plate2, ... 的sheet）
wb.read_from_excel("input_plates.xlsx")

# 3. 获取 plate1 的编号 → 内容 映射
plate1 = wb.plates.get("plate1")
if plate1:
    numbered = plate1.get_numbered_values()
    print("前10个编号位置内容：")
    for i in range(1, 11):
        print(f"Position {i}: {numbered.get(i)}")

    # 4. 示例：通过编号查内容
    value_at_25 = plate1.get_value_by_number(25)
    print(f"编号25的内容是：{value_at_25}")

    # 5. 示例：通过内容查编号
    if value_at_25:
        positions = plate1.get_number_by_value(value_at_25)
        print(f"内容“{value_at_25}”出现在编号位置：{positions}")
else:
    print("未找到 plate1 数据。")

# 6. 将所有 plate 写回新的 Excel 文件
wb.write_to_excel("output_plates.xlsx")

# 7. 将list写为plate
data = [f"Sample{i+1}" for i in range(45)]  # 测试用不满96长度的数据
pb = PlateWorkbook.from_list(data, plate_name="plate1")
pb.write_to_excel("demo_plate.xlsx")

# ===========================匹配高亮===========================

# 假设你的workbook变量名为pw，原始文件路径为plate_file
pw.highlight_and_save(plate_file, plate_highlight_dict)

# 1. 读取Excel，生成PlateWorkbook
pw = PlateWorkbook()
pw.read_from_excel('your_plate_file.xlsx')

# 2. 匹配得到要高亮的孔号，形式为dict: plate_name -> set(编号)
highlight_dict = {
    'plate1': {1, 2, 3},
    'plate2': {12, 24, 36}
}

# 3. 输出高亮结果
highlighted_file = pw.highlight_and_save('your_plate_file.xlsx', highlight_dict)
print(f'高亮文件已保存为: {highlighted_file}')


'''