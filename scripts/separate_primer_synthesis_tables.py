import tkinter as tk
from tkinter import filedialog, messagebox
import os
import openpyxl
from pathlib import Path


def extract_subtables(filepath):
    # 获取文件名（不带扩展名）
    input_file = Path(filepath)
    file_name = input_file.stem
    

    # 打开 Excel
    wb = openpyxl.load_workbook(filepath, data_only=True)

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]

        # 遍历寻找两个 9x13 子表
        # 按规则：表头第一行 "1~12"，第一列 "A~H"
        start_cells = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=2, max_col=ws.max_column):
            for cell in row:
                if str(cell.value) == "1":  # 表头第一列找到 "1"
                    # 检查接下来的 12 列是否为 1~12
                    values = [ws.cell(cell.row, cell.column + i).value for i in range(12)]
                    if values == list(range(1, 13)):
                        # 检查下面 8 行是否 A~H
                        col_A = [ws.cell(cell.row + 1 + i, cell.column - 1).value for i in range(8)]
                        if col_A == list("ABCDEFGH"):
                            start_cells.append((cell.row, cell.column - 1))

        if len(start_cells) != 2:
            print(f"⚠️ {sheetname} 中未找到两个子表，跳过。")
            continue

        for idx, (start_row, start_col) in enumerate(start_cells):
            # 取出 9x13 区域（含表头）
            table_data = []
            for r in range(start_row, start_row + 9):
                row_data = []
                for c in range(start_col, start_col + 13):
                    row_data.append(ws.cell(r, c).value)
                table_data.append(row_data)

            # 创建新的 workbook 保存
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active

            for r, row_data in enumerate(table_data, start=1):
                for c, value in enumerate(row_data, start=1):
                    new_ws.cell(r, c, value)

            # 判断命名规则
            if any("_" in str(v) or "-" in str(v) for row in table_data for v in row if v is not None):
                suffix = "plate"
            else:
                suffix = "conc"
            
            output_file = input_file.with_stem(f"{sheetname}_{suffix}_{file_name}")
            new_wb.save(output_file)
            print(f"✅ 导出: {output_file}")


def main():
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口

    filepath = filedialog.askopenfilename(
        title="选择一个 Excel 文件",
        filetypes=[("Excel 文件", "*.xlsx")]
    )

    if not filepath:
        messagebox.showwarning("提示", "未选择文件")
        return

    extract_subtables(filepath)

    messagebox.showinfo("完成", "子表提取完成！")


if __name__ == "__main__":
    main()
