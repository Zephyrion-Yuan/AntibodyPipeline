from core.struc_plates_96 import PlateWorkbook

import pandas as pd
import re
import tkinter as tk
from tkinter import filedialog, messagebox
import os
from typing import Tuple
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys
from pathlib import Path

def natural_key(string):
    # 用tuple替代list, tuple可hash
    return tuple(int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', str(string)))


# 映射96孔板字典
# 定义字母序列（A-H）
letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

# 初始化正向映射（数字→字符串）和反向映射（字符串→数字）
n2s = {}
s2n = {}

# 生成映射关系
for num_suffix in range(1, 13):  # 数字部分：1到12
    for idx, letter in enumerate(letters):  # 字母部分：A到H
        # 计算对应的数字（1-96）
        number = (num_suffix - 1) * 8 + (idx + 1)
        # 生成对应的字符串（如"A1"、"B3"等）
        string = f"{letter}{num_suffix}"
        # 填充映射字典
        n2s[number] = string
        s2n[string] = number

# 组合为双向字典（可通过两种方式访问）
bi_dict = {
    'n2s': n2s,
    's2n': s2n
}

def read_xlsx_column(file_path):
    data = []
    try:
        # 读取完整表格（无标题行）
        df = pd.read_excel(file_path, header=None)
        # 提取第2~13列（Excel列B到M），第2~9行（Excel行2到9）
        sub_df = df.iloc[1:9, 1:13]  # pandas索引从0开始
        # 按列优先展开为列表
        data = sub_df.values.T.flatten().tolist()
    except FileNotFoundError:
        print(f"错误：文件 '{file_path}' 不存在")
    except Exception as e:
        print(f"读取文件时发生错误: {e}")
    return data


def calculate_conc_arm(plate_path: str, conc_path: str, des_conc:float, volume:float, parent=None) -> Tuple[PlateWorkbook, pd.DataFrame]:
    """
    根据提供的plate文件和浓度文件，计算质粒浓度并生成结果
    
    参数:
    plate_path (str): 包含plate数据的Excel文件路径
    conc_path (str): 浓度标准曲线参数文件路径

    返回:
    Tuple[PlateWorkbook, pd.DataFrame]: 处理后的PlateWorkbook对象和输出DataFrame
    """
    wb = PlateWorkbook()
    wb.read_from_excel(plate_path)

    # 读取浓度文件（例如txt或csv）
    data_list = read_xlsx_column(conc_path)
    if not data_list:
        raise ValueError("浓度文件为空或读取失败")

    df_conc = pd.DataFrame(data_list, columns=['Concerntration'])

    # 添加抗体名称
    first_plate_key = next(iter(wb.plates.keys()))  # 取第一个键（例如 '1'）
    df_conc['Arm'] = wb.plates[first_plate_key].data.iloc[1:,1:].values.flatten(order='F').tolist()
    

    # 计算所需体积 v = adjusted_conc * volume / conc - adjusted_conc，并保留两位小数
    # 定义处理单个数值的函数
    def calculate_volume(x, des_conc, volume):
        try:
            # 核心计算
            denominator = x - des_conc
            if denominator == 0:  # 避免除以0s
                result = 0.0
            else:
                result = (des_conc * volume) / denominator
            
            # 处理无穷大/负无穷/空值
            if pd.isna(result) or result in (float('inf'), float('-inf')):
                result = 0.0
            # 限制最小值为0，保留3位小数
            result = max(result, 0.0)  # 替代 clip(lower=0)
            result = round(result, 3)  # 保留3位小数
            return result
        except:
            # 捕获其他异常（如x本身是NaN）
            return 0.0
    
    df_conc['Volume'] = df_conc['Concerntration'].apply(calculate_volume, args=(des_conc, volume))
    # df_conc['Volume'] = df_conc['Concerntration'].rdiv(m).replace([float('inf'), float('-inf')], 0).fillna(0).clip(lower=0).round(3)
    # 调整表格顺序
    output_frame = df_conc[['Arm', 'Concerntration', 'Volume']]
    warn_high = []
    warn_low = []
    # 使用iterrows()替代手动索引
    des_conc_trans = des_conc
    for idx, row in df_conc.iterrows():
        plas_name = row['Arm']
        if not plas_name.strip():  # 使用strip()处理空字符串或空格
            continue
        if row['Concerntration'] < des_conc_trans:
            warn_low.append(idx)
        elif row['Volume'] < 0.01:
            warn_high.append(idx)

    output_frame = output_frame[(output_frame['Arm'].notna()) & (output_frame['Arm'] != '')]

    output_path = Path(conc_path)
    file_name = output_path.stem
    output_file = output_path.with_stem(f"{file_name}_primer_conc")
    output_frame.to_excel(output_file, index=False)

    wbs = load_workbook(output_file)

    if warn_high or warn_low:
        ws = wbs.active
        # 定义填充样式
        fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # 红色
        fill_blue = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # 蓝色

        # 注意：Excel中索引从1开始，且第一行是表头
        for row_idx in warn_high:
            excel_row = row_idx + 2  # 加2是因为Excel第一行为表头，pandas从0开始
            ws[f"A{excel_row}"].fill = fill_red

        for row_idx in warn_low:
            excel_row = row_idx + 2
            ws[f"A{excel_row}"].fill = fill_blue
        wbs.save(output_file)

    if warn_high and warn_low:
        messagebox.showinfo("警告", "存在样本浓度过高和过低，请调整", parent=parent)
        sys.exit(1)
    elif warn_high:
        messagebox.showinfo("警告", "存在样本浓度过高，请调整", parent=parent)
        sys.exit(1)
    elif warn_low:
        messagebox.showinfo("警告", "存在样本浓度过低，请检查", parent=parent)
        sys.exit(1)
    
    return wb, output_frame


def output_worklist(path, frame, bi_dict):
    # 首先过滤掉Arm列为空和''的行
    frame = frame[(frame['Arm'].notna()) & (frame['Arm'] != '')]

    # 检查是否存在重复值
    duplicates = frame['Arm'].duplicated()

    if duplicates.any():
        # 找出重复的值及其位置
        duplicate_values = frame.loc[duplicates, 'Arm'].unique()
        duplicate_values_str = '\n - '.join(map(str, duplicate_values))
        # 抛出异常并退出
        error_msg = f"检测到Arm列存在重复值！重复值为: \n - {duplicate_values_str}"
        messagebox.showerror("错误", error_msg)
        raise ValueError("样本出现重复")

    columns = [
        "Source_label", "Source_position", "Destination_position", "Destination_label",  "Volume", "Arm"
    ]
    plas_df = pd.DataFrame(columns=columns)

    plas_df = pd.DataFrame({
        'Source_label': ['1'] * len(frame),
        'Source_position': list(range(1, len(frame) + 1)),
        'Destination_position': list(range(1, len(frame) + 1)),
        'Destination_label': ['cherry'] * len(frame),
        'Volume': frame['Volume'],
        'Arm': frame['Arm'],
        
    })

    columns_xls = [
        "Sample_labware",	"Sample_position",	"Sample_Volume",	"Destination_labware",	"Destination_position",	"Sample_ID"
    ]

    plas_df_xls = pd.DataFrame(columns=columns_xls)
    plas_df_xls = pd.DataFrame({
        'Sample_labware': "sample1",
        'Sample_position': [bi_dict['n2s'][i] for i in range(1, len(frame) + 1)],
        'Sample_Volume': frame['Volume'],
        'Destination_labware': "cherry",
        'Destination_position': [bi_dict['n2s'][i] for i in range(1, len(frame) + 1)],
        'Sample_ID': frame['Arm'],
    })
    
    output_path = Path(path)
    plas_df.to_csv(output_path.with_name(f"{output_path.stem}_all.csv"), index=False)
    plas_df_xls.to_excel(output_path.with_name(f"{output_path.stem}_all.xlsx"),index=False)
    return


def get_user_inputs():

    # 回传结果
    result = {}

    def choose_file(var):
        path = filedialog.askopenfilename()
        if path:
            var.set(path)

    def on_ok():
        # 检查路径
        paths = []
        for var in path_vars:
            value = var.get().strip()
            if not value:
                messagebox.showerror("错误", "所有路径不能为空")
                return
            paths.append(value)
        # 检查数字
        floats = []
        for entry, var in zip(num_entries, num_vars):
            value = var.get().strip()
            try:
                num = float(value)
                floats.append(num)
            except Exception:
                messagebox.showerror("错误", f"请输入有效数字：{entry['label_text']}")
                return
            
        # 返回数据
        nonlocal result
        result = {
            "plate_path": paths[0],
            "conc_path": paths[1],
            "conc": floats[0],
            "volume": floats[1],
            "parent": root,
        }
        # root.destroy()
        root.withdraw()
        root.quit()

    def on_cancel():
        root.destroy()
        return

    root = tk.Toplevel()
    root.title("输入参数")
    root.resizable(False, False)

    # 设置为顶层窗口并获得焦点
    root.transient(root.master)
    root.grab_set()

    frm = tk.Frame(root)
    frm.pack(padx=16, pady=16)

    # 路径输入
    path_labels = ["样本布局", "返测浓度"]
    path_vars = [tk.StringVar() for _ in range(2)]
    for i, label in enumerate(path_labels):
        row = tk.Frame(frm)
        row.pack(fill='x', pady=3)
        tk.Label(row, text=label, width=13, anchor='e').pack(side='left')
        entry = tk.Entry(row, textvariable=path_vars[i], width=40)
        entry.pack(side='left', padx=5)
        btn = tk.Button(row, text="选择", command=lambda v=path_vars[i]: choose_file(v))
        btn.pack(side='left')

    # 数值输入
    num_labels = [
        ("目标浓度(ng/ul)", ""),
        ("加水体积(ul)", ""),
    ]
    num_vars = [tk.StringVar(value=str(val)) for label, val in num_labels]
    num_entries = []
    for i, (label, val) in enumerate(num_labels):
        row = tk.Frame(frm)
        row.pack(fill='x', pady=2)
        lbl = tk.Label(row, text=label, width=13, anchor='e')
        lbl.pack(side='left')
        entry = tk.Entry(row, textvariable=num_vars[i], width=15, justify='right')
        entry.icursor('end')
        entry.pack(side='left')
        entry.label_text = label
        num_entries.append(entry)
        # 只允许数字和小数点
        def only_float_input(text):
            if text == "":
                return True
            try:
                float(text)
                return True
            except:
                return False
        entry.config(validate="key",
                     validatecommand=(root.register(only_float_input), "%P"))

    # 按钮
    btns = tk.Frame(frm)
    btns.pack(pady=(16,0))
    tk.Button(btns, text="确定", command=on_ok, width=12).pack(side='left', padx=8)
    tk.Button(btns, text="取消", command=on_cancel, width=12).pack(side='left', padx=8)

    root.mainloop()
    return result


def main():

    inputs = get_user_inputs()

    plate_path = inputs['plate_path']
    conc_path = inputs['conc_path']

    filename = os.path.basename(plate_path)

    conc, volume = inputs['conc'], inputs['volume']
    parent = inputs['parent']

    _, output_frame = calculate_conc_arm(plate_path, conc_path, conc, volume, parent=parent)

    output_worklist(plate_path, output_frame, bi_dict)

    parent.destroy()


    # 构建弹窗统计信息
    result_msg = (
        f"处理完成！\n\n"
    )

    messagebox.showinfo("完成", result_msg)

    return

if __name__ == "__main__":
    main()
    