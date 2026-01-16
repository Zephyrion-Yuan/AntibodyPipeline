import tkinter as tk
from tkinter import filedialog, messagebox
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import InvalidFileException

def select_files():
    """选择txt文件和xlsx文件"""
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口
    
    # 选择txt文件
    txt_file = filedialog.askopenfilename(
        title="选择包含样本的TXT文件",
        filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
    )
    
    if not txt_file:
        messagebox.showwarning("警告", "未选择TXT文件")
        return None, None
    
    # 选择xlsx文件
    xlsx_file = filedialog.askopenfilename(
        title="选择要检索的XLSX文件",
        filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
    )
    
    if not xlsx_file:
        messagebox.showwarning("警告", "未选择XLSX文件")
        return None, None
        
    return txt_file, xlsx_file

def normalize_sample(value):
    """标准化样本值，去除空白并统一字符串格式。"""
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def read_samples(txt_file):
    """从txt文件中读取样本，每行一个样本。"""
    try:
        with open(txt_file, 'r', encoding='utf-8') as f:
            # 使用集合去重并加速后续匹配
            samples = {
                normalized
                for line in f
                if (normalized := normalize_sample(line))
            }
        return samples
    except Exception as e:
        messagebox.showerror("错误", f"读取TXT文件时出错: {str(e)}")
        return None

def highlight_matches(xlsx_file, samples):
    """在xlsx文件中高亮匹配的样本。"""
    try:
        # 加载工作簿
        wb = load_workbook(xlsx_file)

        # 创建黄色填充样式
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"正在处理工作表: {sheet_name}")

            # 遍历所有单元格，使用集合加速匹配
            for row in sheet.iter_rows():
                for cell in row:
                    cell_value = normalize_sample(cell.value)
                    if cell_value and cell_value in samples:
                        print(cell_value)
                        cell.fill = yellow_fill
        
        # 生成输出文件名和路径
        xlsx_dir = os.path.dirname(xlsx_file)
        xlsx_name = os.path.basename(xlsx_file)
        output_file = os.path.join(xlsx_dir, f"highlighted_{xlsx_name}")
        
        # 保存高亮后的文件
        wb.save(output_file)
        return output_file
        
    except InvalidFileException:
        messagebox.showerror("错误", "选择的文件不是有效的XLSX文件")
        return None
    except Exception as e:
        messagebox.showerror("错误", f"处理XLSX文件时出错: {str(e)}")
        return None

def main():
    # 选择文件
    txt_file, xlsx_file = select_files()
    if not txt_file or not xlsx_file:
        return
        
    # 读取样本
    samples = read_samples(txt_file)
    if not samples:
        messagebox.showwarning("警告", "未从TXT文件中读取到任何样本")
        return

    print(f"已读取 {len(samples)} 个样本")
    
    # 高亮匹配
    output_file = highlight_matches(xlsx_file, samples)
    
    if output_file:
        messagebox.showinfo("完成", f"处理完成！\n文件已保存至: {output_file}")
    else:
        messagebox.showerror("失败", "处理过程中出现错误，未能生成输出文件")

if __name__ == "__main__":
    main()
