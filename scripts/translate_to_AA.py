# -*- coding: utf-8 -*-
import re
import sys
from tkinter import Tk, messagebox, filedialog

from openpyxl import load_workbook

# 标准遗传密码（DNA，T 而非 U）
CODON_TABLE = {
    # 苯丙氨酸/亮氨酸等
    "TTT":"F","TTC":"F","TTA":"L","TTG":"L",
    "CTT":"L","CTC":"L","CTA":"L","CTG":"L",
    "ATT":"I","ATC":"I","ATA":"I","ATG":"M",
    "GTT":"V","GTC":"V","GTA":"V","GTG":"V",

    "TCT":"S","TCC":"S","TCA":"S","TCG":"S",
    "CCT":"P","CCC":"P","CCA":"P","CCG":"P",
    "ACT":"T","ACC":"T","ACA":"T","ACG":"T",
    "GCT":"A","GCC":"A","GCA":"A","GCG":"A",

    "TAT":"Y","TAC":"Y","TAA":"*","TAG":"*",
    "CAT":"H","CAC":"H","CAA":"Q","CAG":"Q",
    "AAT":"N","AAC":"N","AAA":"K","AAG":"K",
    "GAT":"D","GAC":"D","GAA":"E","GAG":"E",

    "TGT":"C","TGC":"C","TGA":"*","TGG":"W",
    "CGT":"R","CGC":"R","CGA":"R","CGG":"R",
    "AGT":"S","AGC":"S","AGA":"R","AGG":"R",
    "GGT":"G","GGC":"G","GGA":"G","GGG":"G",
}

def clean_and_to_dna(seq: str) -> str:
    """
    1) 去掉小写字母 [a-z]
    2) 保留并大写 A/T/G/C/U，其它字符丢弃
    3) 将 U 转为 T（把 RNA 当作 DNA 处理）
    """
    if seq is None:
        return ""
    # 去小写
    s = re.sub(r"[a-z]+", "", str(seq))
    # 统一大写
    s = s.upper()
    # 仅保留核酸字符
    s = re.sub(r"[^ATGCU]", "", s)
    # U -> T
    s = s.replace("U", "T")
    return s

def translate_dna(dna: str) -> str:
    """
    将 DNA 序列按密码子翻译为单字母氨基酸序列。
    - 长度非 3 的倍数，末尾余数丢弃
    - 未知或含 N 的密码子用 'X'
    - 终止密码子写入 '*'
    """
    aa = []
    n = len(dna) // 3 * 3
    for i in range(0, n, 3):
        codon = dna[i:i+3]
        aa.append(CODON_TABLE.get(codon, "X"))
    return "".join(aa)

def main():
    # 选择文件
    root = Tk()
    root.withdraw()  # 不显示主窗口
    path = filedialog.askopenfilename(
        title="选择 Excel 文件",
        filetypes=[("Excel 工作簿", "*.xlsx")]
    )
    if not path:
        messagebox.showwarning("已取消", "未选择文件。")
        return

    try:
        wb = load_workbook(path)
    except Exception as e:
        messagebox.showerror("打开失败", f"无法打开文件：\n{e}")
        return

    ws = wb.active  # 处理活动工作表
    changed_rows = 0

    # 遍历所有已用行：读取 B 列，写入 C 列
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value  # B列
        if cell_val is None or str(cell_val).strip() == "":
            # 如果 B 列为空，则 C 列也清空（可选）
            # ws.cell(row=row, column=3).value = ""
            continue

        dna = clean_and_to_dna(cell_val)
        aa = translate_dna(dna)
        ws.cell(row=row, column=3).value = aa
        changed_rows += 1

    try:
        wb.save(path)
    except Exception as e:
        messagebox.showerror("保存失败", f"写回原文件失败：\n{e}")
        return
    finally:
        wb.close()

    messagebox.showinfo("完成", f"已处理并写入第3列（C 列）。\n共处理 {changed_rows} 行。\n文件：{path}")



if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        # 确保异常有提示
        messagebox.showerror("运行出错", f"{e}")
        sys.exit(1)
