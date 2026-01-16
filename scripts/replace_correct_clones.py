import os
import tkinter as tk
from tkinter import filedialog, messagebox

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def select_multiple_xlsx():
    """
    允许用户多次选择 xlsx 文件（可来自不同目录），直到点击取消或选择“不再继续”。
    返回去重后的文件路径列表。
    """
    paths = []
    while True:
        files = filedialog.askopenfilenames(
            title="选择需要修改的 xlsx 文件（可多次选择不同目录）",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not files:
            # 用户直接取消：如果还没选过任何文件，就直接结束
            break

        paths.extend(files)

        # 询问是否继续选择
        cont = messagebox.askyesno("继续选择？", "是否继续从其他目录选择更多的 xlsx 文件？")
        if not cont:
            break

    # 去重并保持顺序
    unique_paths = list(dict.fromkeys(paths))
    return unique_paths


def process_files(csv_path, xlsx_paths):
    """
    处理逻辑：
    - 读取 csv（带表头）
    - 遍历每一行（跳过表头）的前两列：
        col1, col2 = 第一列文本, 第二列文本
    - 在所有 xlsx 中查找与 col2 匹配的单元格（按字符串比较，去掉首尾空格）
      若找到 → 单元格值改为 f"{col1}-0"，背景设为浅蓝色
      若未在任何文件中找到 → 记录 col2
    - 保存所有 xlsx 文件
    返回 (修改的单元格数量, 未找到文本列表)
    """
    # 读取 CSV
    df = pd.read_csv(csv_path)  # 如有编码问题可加 encoding="utf-8"

    if df.shape[1] < 2:
        raise ValueError("CSV 至少需要两列数据")

    # 提前把前两列取出来，便于处理
    col1_values = df.iloc[:, 0].tolist()
    col2_values = df.iloc[:, 1].tolist()

    # 预先加载所有 xlsx 工作簿
    workbooks = []  # list of (path, workbook)
    for path in xlsx_paths:
        wb = load_workbook(path)
        workbooks.append((path, wb))

    # 设置浅蓝色填充
    light_blue_fill = PatternFill(
        start_color="ADD8E6",  # 浅蓝色
        end_color="ADD8E6",
        fill_type="solid"
    )

    total_modified_cells = 0
    found_texts = []
    not_found_texts = []  # 记录未找到的 col2 文本（去重但保持顺序）
    seen_not_found = set()

    # 遍历 CSV 中每一行的数据
    # 跳过表头已经在 read_csv 时自动处理，因此直接按行遍历即可
    for col1, col2 in zip(col1_values, col2_values):
        if pd.isna(col2):
            continue  # 第二列为空就跳过

        search_text = str(col2).strip()
        replacement_text = f"{col1}-9"

        found_any = False

        # 在所有 workbook 的所有 sheet 的所有单元格中查找
        for path, wb in workbooks:
            for ws in wb.worksheets:
                # 遍历所有单元格
                for row in ws.iter_rows():
                    for cell in row:
                        cell_value = cell.value
                        if cell_value is None:
                            continue

                        # 统一按字符串比较，并 strip 首尾空格
                        if str(cell_value).strip() == search_text:
                            # 修改单元格内容
                            cell.value = replacement_text
                            found_texts.append(replacement_text)
                            # 修改背景色为浅蓝色
                            cell.fill = light_blue_fill

                            total_modified_cells += 1
                            found_any = True

        # 如果在所有 xlsx 中都没找到
        if not found_any:
            if search_text not in seen_not_found:
                seen_not_found.add(search_text)
                not_found_texts.append(search_text)

    # # 处理完后保存所有工作簿（覆盖原文件）
    for path, wb in workbooks:
    #       wb.save(path)
        base, ext = os.path.splitext(path)
        new_path = base + "_modified" + ext
        wb.save(new_path)
    
    # 输出未匹配的样本
    base, ext = os.path.splitext(csv_path)
    path_not_found = base + "_not_found.txt"  # 改为txt后缀

    # 将列表内容逐行写入txt文件
    with open(path_not_found, 'w', encoding='utf-8') as f:
        for text in not_found_texts:
            f.write(f"{text}\n")

    # 输出匹配的样本
    path_found = base + "_found.txt"  # 改为txt后缀
    with open(path_found, 'w', encoding='utf-8') as f:
        for text in found_texts:
            f.write(f"{text}\n")

    return total_modified_cells, not_found_texts


def main():
    root = tk.Toplevel()
    root.withdraw()  # 隐藏主窗口，仅使用文件对话框和消息框

    # 1) 选择 CSV 文件
    csv_path = filedialog.askopenfilename(
        title="选择 CSV 文件",
        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
    )
    if not csv_path:
        messagebox.showerror("错误", "未选择 CSV 文件。")
        return

    # 2) 选择多个 xlsx 文件（可跨目录多次选择）
    xlsx_paths = select_multiple_xlsx()
    if not xlsx_paths:
        messagebox.showerror("错误", "未选择任何 xlsx 文件。")
        return

    try:
        modified_count, not_found_texts = process_files(csv_path, xlsx_paths)
    except Exception as e:
        messagebox.showerror("错误", f"处理过程中发生错误：\n{e}")
        return

    # 3) 统计结果并弹出提示框
    msg_lines = [
        f"处理完成！",
        f"共修改单元格数量：{modified_count}",
        f"未在任何 xlsx 中找到的文本数量：{len(not_found_texts)}",
        ""
    ]
    if not_found_texts:
        msg_lines.append("以下文本未查找到：")
        msg_lines.extend(not_found_texts)

    final_msg = "\n".join(msg_lines)
    messagebox.showinfo("结果", final_msg)
    root.destroy
    return


if __name__ == "__main__":
    main()
