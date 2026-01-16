from core.struc_plates_96 import PlateWorkbook
from core.read_config_path import get_app_dir, get_meipass_dir, find_config_path

import pandas as pd
import re
import tkinter as tk
from tkinter import filedialog, messagebox
import os
from typing import Tuple
import configparser
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys
from pathlib import Path

def natural_key(string):
    # 用tuple替代list, tuple可hash
    return tuple(int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', str(string)))

def std_curve(c: float, b: float, a: float, scale: float, input_value: float) -> float:
    """
    返回:
    float: 转换后的输出值，保留两位小数
    """
    # 计算标准曲线转换
    denominator = b * scale
    if denominator == 0:
        messagebox.showerror("错误", "b和scale不能为0")
        return
    factor = a / denominator
    
    # 规范化结果，保留两位小数（四舍五入）
    return round((input_value - c) * factor, 6)

def read_asc_column(file_path, sep='\t'):
    data = []
    try:
        with open(file_path, 'r', encoding='utf-16') as file:
            for _ in range(96):
                line = file.readline()
                if not line:
                    break
                # 分割行并提取第1列（索引0）
                columns = line.strip().split(sep)
                if columns:
                    data.append(columns[0])
    except FileNotFoundError:
        print(f"错误：文件 '{file_path}' 不存在")
    except Exception as e:
        print(f"读取文件时发生错误: {e}")
    return data


def calculate_conc_antibody(plate_path: str, conc_path: str, c:float, b:float, a:float, scale:float, des_conc:float, volume:float, parent=None) -> Tuple[PlateWorkbook, pd.DataFrame]:
    """
    根据提供的plate文件和浓度文件，计算质粒浓度并生成结果
    
    参数:
    plate_path (str): 包含plate数据的Excel文件路径
    conc_path (str): 浓度标准曲线参数文件路径

    计算方式
    conc = (read - c) * a / (b * scale)
    volume = des_conc * volume / conc
    
    返回:
    Tuple[PlateWorkbook, pd.DataFrame]: 处理后的PlateWorkbook对象和输出DataFrame
    """
    wb = PlateWorkbook()
    wb.read_from_excel(plate_path)

    # 读取浓度文件（例如txt或csv）
    data_list = read_asc_column(conc_path)
    if not data_list:
        raise ValueError("浓度文件为空或读取失败")

    df_conc = pd.DataFrame(data_list, columns=['BCA'])

    # 使用 std_curve 函数计算 Adjusted concentration
    df_conc['Adjusted concentration'] = df_conc['BCA'].apply(
        lambda x: std_curve(c, b, a, scale, float(x)) if x else None
    )

    # 添加抗体名称
    first_plate_key = next(iter(wb.plates.keys()))  # 取第一个键（例如 '1'）
    df_conc['Antibody'] = wb.plates[first_plate_key].data.iloc[1:,1:].values.flatten(order='F').tolist()

    # 计算所需体积 v = m / adjusted concentration，并保留两位小数
    m = des_conc * volume / 1000.0
    df_conc['Volume'] = df_conc['Adjusted concentration'].rdiv(m).replace([float('inf'), float('-inf')], 0).fillna(0).clip(lower=0).round(3)
    df_conc['Water_volume'] = volume - df_conc['Volume']
    # 调整表格顺序
    output_frame = df_conc[['Antibody', 'BCA', 'Adjusted concentration', 'Volume', 'Water_volume']]
    warn_high = []
    warn_low = []
    # 使用iterrows()替代手动索引
    des_conc_trans = des_conc / 1000.0
    for idx, row in df_conc.iterrows():
        ab_name = row['Antibody']
        if not ab_name.strip():  # 使用strip()处理空字符串或空格
            continue
        if row['Adjusted concentration'] < des_conc_trans:
            warn_low.append(idx)
        elif row['Volume'] < 0.01:
            warn_high.append(idx)

    output_path = os.path.dirname(conc_path)
    output_file = os.path.join(output_path, 'conc_antibody.xlsx')
    output_frame.to_excel(output_file, index=False)

    wbs = load_workbook(output_file)

    if warn_high or warn_low:
        ws = wbs.active
        # 定义填充样式
        fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # 红色
        fill_blue = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # 蓝色

        # 注意：Excel中索引从1开始，且第一行是表头
        for row_idx in warn_high:
            excel_row = row_idx + 2  # 加2是因为Excel第一行为表头，pandas从0开始
            ws[f"A{excel_row}"].fill = fill_red

        for row_idx in warn_low:
            excel_row = row_idx + 2
            ws[f"A{excel_row}"].fill = fill_blue
        wbs.save(output_file)

    if warn_high and warn_low:
        messagebox.showinfo("警告", "存在样本浓度过高和过低，请调整", parent=parent)
        sys.exit(1)
    elif warn_high:
        messagebox.showinfo("警告", "存在样本浓度过高，请调整", parent=parent)
        sys.exit(1)
    elif warn_low:
        messagebox.showinfo("警告", "存在样本浓度过低，请检查", parent=parent)
        sys.exit(1)
    
    return wb, output_frame


def output_worklist(path, frame, filename):
    columns = [
        "Antibody", "Source_label", "Source_position", "Destination_position", "Destination_label",  "Volume",
    ]
    ab_df = pd.DataFrame(columns=columns)
    water_df = pd.DataFrame(columns=columns)

    ab_df = pd.DataFrame({
        'Antibody': frame['Antibody'],
        'Source_label': ['1'] * len(frame),
        'Source_position': list(range(1, len(frame) + 1)),
        'Destination_position': list(range(1, len(frame) + 1)),
        'Destination_label': ['cherry'] * len(frame),
        'Volume': frame['Volume']
    })

    water_df = pd.DataFrame({
        'Antibody': frame['Antibody'],
        'Source_label': ['water'] * len(frame),
        'Source_position': list(range(1, len(frame) + 1)),
        'Destination_position': list(range(1, len(frame) + 1)),
        'Destination_label': ['cherry'] * len(frame),
        'Volume': frame['Water_volume']
    })
    combined_df = pd.concat([ab_df, water_df], axis=0, ignore_index=True)

    file_full_path = os.path.join(os.path.dirname(path), filename)
    combined_df.to_excel(file_full_path, index=False)
    return


def get_user_inputs():

    config_path = "config/std_curve_defaults_antibody.ini"
    # 1. 先加载历史默认值
    defaults = load_defaults(config_path)

    # 回传结果
    result = {}

    def choose_file(var):
        path = filedialog.askopenfilename()
        if path:
            var.set(path)

    def on_ok():
        # 检查路径
        paths = []
        for var in path_vars:
            value = var.get().strip()
            if not value:
                messagebox.showerror("错误", "所有路径不能为空")
                return
            paths.append(value)
        # 检查数字
        floats = []
        for entry, var in zip(num_entries, num_vars):
            value = var.get().strip()
            try:
                num = float(value)
                floats.append(num)
            except Exception:
                messagebox.showerror("错误", f"请输入有效数字：{entry['label_text']}")
                return
            
        # 保存新默认值
        save_defaults(*floats, config_path=config_path)

        # 返回数据
        nonlocal result
        result = {
            "plate_path": paths[0],
            "conc_path": paths[1],
            "a": floats[0],
            "b": floats[1],
            "c": floats[2],
            "scale": floats[3],
            "conc": floats[4],
            "volume": floats[5],
            "parent": root,
        }
        # root.destroy()
        root.withdraw()
        root.quit()

    def on_cancel():
        root.destroy()
        return

    root = tk.Toplevel()
    root.title("输入参数")
    root.resizable(False, False)

    # 设置为顶层窗口并获得焦点
    root.transient(root.master)
    root.grab_set()

    frm = tk.Frame(root)
    frm.pack(padx=16, pady=16)

    # 路径输入
    path_labels = ["样本ELI布局", "吸光度.asc"]
    path_vars = [tk.StringVar() for _ in range(2)]
    for i, label in enumerate(path_labels):
        row = tk.Frame(frm)
        row.pack(fill='x', pady=3)
        tk.Label(row, text=label, width=13, anchor='e').pack(side='left')
        entry = tk.Entry(row, textvariable=path_vars[i], width=40)
        entry.pack(side='left', padx=5)
        btn = tk.Button(row, text="选择", command=lambda v=path_vars[i]: choose_file(v))
        btn.pack(side='left')

    # 说明
    tk.Label(frm, text="标准曲线计算公式：浓度=标曲系数 * (读数 - 标曲偏置) / 标曲分母 / 缩放因子",
             fg="blue", anchor='w').pack(fill='x', pady=(16,3))

    # 数值输入
    num_labels = [
        ("标曲系数", defaults['a']),
        ("标曲分母", defaults['b']),
        ("标曲偏置", defaults['c']),
        ("缩放因子", defaults['scale']),
        ("目标浓度(mg/ml)", defaults['conc']),
        ("目标体积(ul)", defaults['volume']),
    ]
    num_vars = [tk.StringVar(value=str(val)) for label, val in num_labels]
    num_entries = []
    for i, (label, val) in enumerate(num_labels):
        row = tk.Frame(frm)
        row.pack(fill='x', pady=2)
        lbl = tk.Label(row, text=label, width=13, anchor='e')
        lbl.pack(side='left')
        entry = tk.Entry(row, textvariable=num_vars[i], width=15, justify='right')
        entry.icursor('end')
        entry.pack(side='left')
        entry.label_text = label
        num_entries.append(entry)
        # 只允许数字和小数点
        def only_float_input(text):
            if text == "":
                return True
            try:
                float(text)
                return True
            except:
                return False
        entry.config(validate="key",
                     validatecommand=(root.register(only_float_input), "%P"))

    # 按钮
    btns = tk.Frame(frm)
    btns.pack(pady=(16,0))
    tk.Button(btns, text="确定", command=on_ok, width=12).pack(side='left', padx=8)
    tk.Button(btns, text="取消", command=on_cancel, width=12).pack(side='left', padx=8)

    root.mainloop()
    return result


# def load_defaults(config_path="config/std_curve_defaults_antibody.ini"):
#     defaults = {
#         'a': 1.0,
#         'b': 0.75758,
#         'c': 0.1676,
#         'scale': 1.32,
#         'conc': 15.0,
#         'volume': 150.0,
#     }
#     if os.path.exists(config_path):
#         cfg = configparser.ConfigParser()
#         cfg.read(config_path)
#         if 'params' in cfg:
#             for key in defaults:
#                 try:
#                     defaults[key] = float(cfg['params'].get(key, defaults[key]))
#                 except:
#                     pass
#     return defaults


# def save_defaults(a, b, c, scale, conc, volume, config_path="config/std_curve_defaults_antibody.ini"):
#     cfg = configparser.ConfigParser()
#     cfg['params'] = {
#         'a': str(a),
#         'b': str(b),
#         'c': str(c),
#         'scale': str(scale),
#         'conc': str(conc),
#         'volume': str(volume)
#     }
#     with open(config_path, "w") as f:
#         cfg.write(f)

def load_defaults(config_path: str | None = None):
    """加载抗体标准曲线默认参数"""
    if config_path is None:
        config_path = str(find_config_path("std_curve_defaults_antibody.ini"))

    defaults = {
        'a': 1.0,
        'b': 0.75758,
        'c': 0.1676,
        'scale': 1.32,
        'conc': 15.0,
        'volume': 150.0,
    }

    if os.path.exists(config_path):
        cfg = configparser.ConfigParser()
        cfg.read(config_path, encoding="utf-8")
        if 'params' in cfg:
            for key in defaults:
                try:
                    defaults[key] = float(cfg['params'].get(key, defaults[key]))
                except:
                    pass
    return defaults


def save_defaults(a, b, c, scale, conc, volume, config_path: str | None = None):
    """保存抗体标准曲线默认参数"""
    if config_path is None:
        app_dir = get_app_dir()
        cfg_dir = app_dir / "config"
        cfg_dir.mkdir(parents=True, exist_ok=True)
        config_path = str(cfg_dir / "std_curve_defaults_antibody.ini")
    else:
        Path(os.path.dirname(config_path)).mkdir(parents=True, exist_ok=True)

    cfg = configparser.ConfigParser()
    cfg['params'] = {
        'a': str(a),
        'b': str(b),
        'c': str(c),
        'scale': str(scale),
        'conc': str(conc),
        'volume': str(volume),
    }

    with open(config_path, "w", encoding="utf-8") as f:
        cfg.write(f)


def main():
    inputs = get_user_inputs()

    plate_path = inputs['plate_path']
    conc_path = inputs['conc_path']

    a, b, c, scale, conc, volume = inputs['a'], inputs['b'], inputs['c'], inputs['scale'], inputs['conc'], inputs['volume']
    parent = inputs['parent']

    _, output_frame = calculate_conc_antibody(plate_path, conc_path, c, b, a, scale, conc, volume, parent=parent)

    
    
    output_worklist(plate_path, output_frame, filename='cherry_antibody.xlsx')

    parent.destroy()


    # 构建弹窗统计信息
    result_msg = (
        f"处理完成！\n\n"
        f"浓度计算结果: conc_antibody.xlsx \n"
        f"转板worklist: cherry_antibody.xlsx \n"
    )

    messagebox.showinfo("完成", result_msg)

    return

if __name__ == "__main__":
    main()
    